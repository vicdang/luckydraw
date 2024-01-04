# -*- coding: utf-8 -*-
# vim:ts=3:sw=3:expandtabs
"""
---------------------------
Copyright (C) 2023
@Authors: dnnvu
@Date: 28-Nov-23
@Version: 1.0
---------------------------
 Location:
   - src
 Usage example:
   - python script_name.py -i data.xlsx -o data.json
"""

# Import
import argparse
import json
import logging
from openpyxl import load_workbook
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Logger
logger = logging.getLogger(__name__)

class ExcelToJsonConverter:
    """Converts Excel to JSON using openpyxl"""
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.data = None

    def read_excel(self):
        """Reads Excel file"""
        workbook = load_workbook(filename=self.excel_file_path, read_only=True, data_only=True, keep_links=False)

        # Assuming data is in the first sheet (index 0), change this if needed
        sheet = workbook.worksheets[0]

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        logger.debug(data)

        headers = data[0]
        self.data = [{headers[i]: str(row[i]) for i in range(len(headers))} for row in data[1:]]

    def convert_to_json(self, json_file_path):
        """Converts data to JSON and saves it"""
        if self.data is None:
            raise ValueError("No data to convert. Read the Excel file first.")

        # Convert list of dictionaries to dictionary using 'uid' as key
        json_data = {item.pop('uid'): item for item in self.data}

        # Save data as JSON with ensure_ascii=False to preserve non-ASCII characters
        with open(json_file_path, 'w', encoding='utf-8') as file:
            json.dump(json_data, file, indent=2, ensure_ascii=False)

def setup_logging():
    """Sets up logging"""
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')
    return logging.getLogger()

def parse_arguments():
    """Parses command-line arguments"""
    parser = argparse.ArgumentParser(description='Convert Excel file to JSON with specified format')
    parser.add_argument('-i', '--input-file', default='./data.xlsx', help='Input Excel file path')
    parser.add_argument('-o', '--output-file', default='../data.json', help='Output JSON file path')
    parser.add_argument('-d', '--debug', default=False, action='store_true', help='Enable debugging')
    return parser.parse_args()

def main():
    """Main function"""
    logger = setup_logging()
    logger.info('Converting Excel to JSON...')
    args = parse_arguments()
    if args.debug:
        logger.setLevel(logging.DEBUG)
    logger.debug(args)
    converter = ExcelToJsonConverter(args.input_file)
    converter.read_excel()
    converter.convert_to_json(args.output_file)
    
    logger.info('Conversion completed. JSON file saved at %s', args.output_file)

if __name__ == "__main__":
    main()
